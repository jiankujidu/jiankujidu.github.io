# -*- coding: utf-8 -*-
"""
ELISA代测表填写系统 - 后端服务
复制原始Excel模板并填充用户数据，保留所有样式

填写位置分析（基于原始Excel模板）：
- 第3行：B3=姓名, D3=电话(D3:E3合并), G3=单位(G3:I3合并)
- 第4行：B4=快递单号, D4=订单编号(D4:E4合并), G4=业务员(G4:I4合并)
- 第8行：B8=种属(B8:E8合并), G8=样本数量(G8:I8合并)
- 第9行：B9=样本类型(B9:F9合并)
- 第10行：B10=指标名称(B10:I10合并)
- 第11行：A11=实验要求+实验目的(A11:I11合并，需要整体修改)
- 第12行：B12=备注(B12:I12合并)
- 第13行：B13=实验样本信息(B13:I13合并)
- 第14行：B14=样本重复设置(B14:I14合并)
- 第16行：B16=标曲重复设置(B16:I16合并)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

from http.server import HTTPServer, BaseHTTPRequestHandler
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64

# 原始Excel模板路径
TEMPLATE_PATH = r'D:\Desktop\优品Elisa代测表.xlsx'

class ELISAHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        """禁用默认日志"""
        pass
    
    def do_OPTIONS(self):
        """处理CORS预检请求"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def do_POST(self):
        """处理POST请求"""
        if self.path == '/export':
            try:
                # 读取请求体
                content_length = int(self.headers.get('Content-Length', 0))
                body = self.rfile.read(content_length)
                data = json.loads(body.decode('utf-8'))
                
                print(f'收到导出请求，数据: {data.get("name", "")}')
                
                # 填充Excel
                excel_data = fill_excel(data)
                
                # 返回base64编码的Excel文件
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                
                response = json.dumps({
                    'success': True,
                    'data': base64.b64encode(excel_data).decode('utf-8'),
                    'filename': '优品Elisa代测表_已填写.xlsx'
                }, ensure_ascii=False)
                self.wfile.write(response.encode('utf-8'))
                print('导出成功')
            except Exception as e:
                print(f'导出错误: {e}')
                import traceback
                traceback.print_exc()
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode('utf-8'))
        else:
            self.send_error(404)
    
    def do_GET(self):
        """处理GET请求"""
        if self.path == '/health':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(b'{"status": "ok"}')
        else:
            self.send_error(404)


def fill_excel(data):
    """
    填充Excel数据，保留原始样式
    """
    # 加载原始模板
    wb = load_workbook(TEMPLATE_PATH)
    ws1 = wb['样本检测要求信息表']
    
    # ===== 第3行：客户基本信息 =====
    ws1['B3'] = data.get('name', '')  # 姓名
    ws1['D3'] = data.get('phone', '')  # 电话
    ws1['G3'] = data.get('company', '')  # 单位名称
    
    # ===== 第4行：快递和订单信息 =====
    ws1['B4'] = data.get('trackingNo', '')  # 快递单号
    ws1['D4'] = data.get('orderNo', '')  # 订单编号
    ws1['G4'] = data.get('salesperson', '陈美龙')  # 业务员
    
    # ===== 第8行：种属和样本数量 =====
    ws1['B8'] = data.get('species', '')  # 种属
    sample_count = data.get('sampleCount', '')
    ws1['G8'] = int(sample_count) if sample_count else ''  # 样本数量
    
    # ===== 第9行：样本类型 =====
    ws1['B9'] = data.get('sampleType', '')  # 样本类型
    
    # ===== 第10行：指标名称 =====
    ws1['B10'] = data.get('indicator', '')  # 指标名称
    
    # ===== 第11行：实验要求和实验目的 =====
    # A11是合并单元格，内容格式: "实验要求*      实验目的*"
    # 需要保持原有格式，在后面追加用户选择
    test_req = data.get('testRequirement', '')
    test_purpose = data.get('testPurpose', '')
    # 保持原有间距格式
    ws1['A11'] = f'实验要求*{test_req}      实验目的*{test_purpose}'
    
    # ===== 第12行：备注 =====
    ws1['B12'] = data.get('remarks', '')  # 备注
    
    # ===== 第13行：实验样本信息 =====
    ws1['B13'] = data.get('sampleInfo', '')  # 实验样本信息
    
    # ===== 第14行：样本是否做重复 =====
    ws1['B14'] = data.get('sampleRepeat', '无')  # 样本重复设置
    
    # ===== 第16行：标曲是否做重复 =====
    ws1['B16'] = data.get('standardRepeat', '无')  # 标曲重复设置
    
    # ===== Sheet 2: 样本检测信息采集表 =====
    ws2 = wb['样本检测信息采集表']
    
    # 样本数据从第3行开始（第1行是标题，第2行是表头）
    samples = data.get('samples', [])
    for i, sample in enumerate(samples):
        row = i + 3  # 从第3行开始
        if row <= 82:  # 最大80行样本
            ws2.cell(row=row, column=1, value=i + 1)  # 序号
            ws2.cell(row=row, column=2, value=sample.get('stype', ''))  # 样本类型
            ws2.cell(row=row, column=3, value=sample.get('scode', ''))  # 样本编号
            ws2.cell(row=row, column=4, value=sample.get('sgroup', ''))  # 样本分组
            ws2.cell(row=row, column=5, value=sample.get('samount', ''))  # 样本量
            ws2.cell(row=row, column=6, value=sample.get('stemp', ''))  # 保存温度
            ws2.cell(row=row, column=7, value=sample.get('sdate', ''))  # 收集日期
            ws2.cell(row=row, column=8, value=sample.get('snote', ''))  # 备注
            ws2.cell(row=row, column=9, value=sample.get('sreq', ''))  # 处理特殊要求
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def run_server(port=8765):
    """启动HTTP服务器"""
    server = HTTPServer(('127.0.0.1', port), ELISAHandler)
    print(f'ELISA代测表服务运行在 http://127.0.0.1:{port}')
    print('按 Ctrl+C 停止服务')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n服务已停止')
        server.shutdown()


if __name__ == '__main__':
    run_server()
